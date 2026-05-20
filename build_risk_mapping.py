"""Generate the Risk code → pscs_departments mapping workbook for Fatim to fill in.
Two sheets: Mapping (94 risk codes with best-guess suggestions) + PSCS_Reference (all 104 dept rows).
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ----- Mapping rows: (risk_code, suggested_pscs_code, confidence, notes_for_fatim) -----
MAPPING = [
    # Directorates
    ("DIR",      "DIR_HOSP",                  "high",   "Pejabat Pengarah → Hospital Director directorate"),
    ("ADCM",     "DIR_MED",                   "high",   "TP Klinikal (Perubatan) directorate"),
    ("ADCS",     "DIR_SURG",                  "high",   "TP Klinikal (Pembedahan) directorate"),
    ("ADPRO",    "DIR_PROF",                  "high",   "TP Profesional & Operasi directorate"),
    ("ADM",      "DIR_MGMT",                  "high",   "TP Pengurusan directorate"),
    ("ADF",      "DIR_FIN",                   "high",   "TP Kewangan directorate"),
    # Misc admin / governance
    ("PI",       "BAH_GOVERNAN",              "low",    "Pejabat Integriti? Best guess: Division of Governance & Integrity. Confirm."),
    ("JIKK",     "",                          "none",   "Couldn't match. What does JIKK stand for?"),
    ("NUR",     "JAB_KEJURURAWATAN",          "high",   "Nursing Department"),
    ("NURBMU",  "NUR_BED_MGMT",               "high",   "Bed Management Unit (subunit)"),
    ("CORP",    "JAB_KORPORAT",               "high",   "Corporate Communication"),
    ("PUU",     "PEJ_UNDANG_UNDANG",          "high",   "Pejabat Undang-Undang / Legal Office"),
    ("AUX",     "PEJ_POLIS_BANTUAN",          "high",   "Auxiliary Police"),
    ("MDU",     "",                           "none",   "Couldn't match. What does MDU stand for?"),
    ("EXS",     "BAH_PERKHIDMATAN_EKSEKUTIF", "high",   "Division of Executive Services"),
    # Medicine
    ("MED",     "JAB_PERUBATAN",              "high",   "Department of Medicine"),
    ("MEDDERM", "",                           "none",   "Dermatology subunit — not in pscs_departments yet; add new sub-unit row?"),
    ("MEDENDO", "",                           "none",   "Endocrinology subunit — not in pscs_departments yet"),
    ("MEDGER",  "",                           "none",   "Geriatrics subunit — not in pscs_departments yet"),
    ("MEDHAEM", "",                           "none",   "Haematology subunit — not in pscs_departments yet"),
    ("MEDID",   "",                           "none",   "Infectious Diseases subunit — not in pscs_departments yet"),
    ("MEDNEU",  "",                           "none",   "Neurology subunit — not in pscs_departments yet"),
    ("MEDPAL",  "",                           "none",   "Palliative subunit — not in pscs_departments yet"),
    ("MEDRESP", "PUS_RESPIRATORI",            "medium", "Could be the Respiratory Service Centre OR a Medicine subunit. Pick one."),
    ("MEDRHEUM","",                           "none",   "Rheumatology subunit — not in pscs_departments yet"),
    ("ED",      "JAB_PERUBATAN_KECEMASAN",    "high",   "Emergency Medicine"),
    # Rehab
    ("RHB",     "JAB_PERUBATAN_PEMULIHAN",    "high",   "Rehabilitation Medicine"),
    ("RHBOCCT", "",                           "none",   "Occupational Therapy — not in pscs_departments yet"),
    ("RHBPHY",  "",                           "none",   "Physiotherapy — not in pscs_departments yet"),
    ("RHBPO",   "",                           "none",   "Prosthetics & Orthotics — not in pscs_departments yet"),
    ("RHBSLT",  "",                           "none",   "Speech-Language Therapy (Rehab) — not in pscs_departments yet"),
    # Psychiatry
    ("PSY",     "JAB_PSIKIATRI",              "high",   "Psychiatry"),
    ("PSYCP",   "",                           "none",   "Child Psychiatry subunit? — not in pscs_departments yet"),
    # Medical (other)
    ("MLE",     "JAB_ETIKA_PERUBATAN",        "high",   "Medical Ethics & Medical Laws"),
    ("PCM",     "JAB_PERUBATAN_PRIMER",       "high",   "Primary Care Medicine"),
    ("PAED",    "JAB_PEDIATRIK",              "high",   "Paediatrics"),
    ("PAEDNICU","PED_NICU",                   "high",   "NICU subunit of Paediatrics"),
    ("PAEDPICU","PED_PICU",                   "high",   "PICU subunit of Paediatrics"),
    ("PHM",     "JAB_KESIHATAN_AWAM",         "medium", "Public Health Medicine? — confirm"),
    ("NEPH",    "PUS_NEFROLOGI",              "high",   "Nephrology Service Centre"),
    ("ONCO",    "",                           "none",   "Oncology — no PUS_ONKOLOGI in pscs_departments yet"),
    ("CARD",    "PUS_KARDIOLOGI",             "high",   "Cardiology Service Centre"),
    ("DCS",     "PUS_RAWATAN_HARIAN",         "high",   "Day Care Service Centre"),
    ("GHS",     "PUS_GASTRO_HEPATOLOGI",      "high",   "Gastroenterology & Hepatology Service Centre"),
    ("IPC",     "UNIT_INFEKSI",               "high",   "Infection Prevention & Control"),
    # Surgical
    ("SURG",    "JAB_PEMBEDAHAN",             "high",   "Department of Surgery"),
    ("CVTS",    "JAB_CVTS",                   "high",   "Cardiovascular & Thoracic Surgery"),
    ("ORTHO",   "JAB_ORTHO",                  "high",   "Orthopaedics & Traumatology"),
    ("OPH",     "JAB_OFTALMOLOGI",            "high",   "Ophthalmology"),
    ("OPHOPTOM","",                           "none",   "Optometry subunit — not in pscs_departments yet"),
    ("ORL",     "JAB_ORL",                    "high",   "Otorhinolaryngology – Head & Neck Surgery"),
    ("ORLAUD",  "",                           "none",   "Audiology subunit — not in pscs_departments yet"),
    ("ORLSLT",  "",                           "none",   "Speech-Language Therapy (ORL) subunit — not in pscs_departments yet"),
    ("ANA",     "JAB_ANESTESIOLOGI",          "high",   "Anaesthesiology & Intensive Care"),
    ("ANAAPS",  "",                           "none",   "Acute Pain Service subunit — not in pscs_departments yet"),
    ("ANACICU", "ANES_CICU_PPUITM",           "high",   "Cardiothoracic ICU (PPUiTM)"),
    ("ANAICU",  "ANES_ICU_HASA",              "medium", "Two options: ANES_ICU_HASA or ANES_ICU_PPUITM. Pick one (or split into two risk codes)."),
    ("OG",      "JAB_OBG",                    "high",   "Obstetrics & Gynaecology"),
    ("FOR",     "",                           "none",   "Forensic Medicine — not in pscs_departments yet"),
    ("DENT",    "",                           "none",   "Dental — not in pscs_departments yet"),
    ("PRS",     "PUS_PLASTIK",                "high",   "Plastic Surgery Service Centre"),
    ("OT",      "PUS_OT",                     "high",   "Operation Theatre Service Centre"),
    # Diagnostics / Professional
    ("CDL",     "JAB_MAKMAL",                 "high",   "Clinical Diagnostic Laboratory"),
    ("CDLAP",   "",                           "none",   "Anatomical Pathology subunit — not in pscs_departments yet"),
    ("CDLCP",   "",                           "none",   "Chemical Pathology subunit — not in pscs_departments yet"),
    ("CDLHTM",  "",                           "none",   "Haematology Pathology subunit — not in pscs_departments yet"),
    ("CDLMMP",  "",                           "none",   "Microbiology & Mycology Pathology subunit — not in pscs_departments yet"),
    ("CDLGP",   "",                           "none",   "General Pathology subunit — not in pscs_departments yet"),
    ("RAD",     "JAB_RADIOLOGI",              "high",   "Radiology"),
    ("QIPS",    "JAB_QPS",                    "high",   "Quality Improvement & Patient Safety (currently Department of Risk Management, Compliance & Quality)"),
    ("PHR",     "JAB_FARMASI",                "high",   "Pharmacy"),
    ("MR",      "JAB_MAKLUMAT_PESAKIT",       "high",   "Medical Records / Patient Information"),
    ("DIET",    "JAB_DIETETIK",               "high",   "Dietetics & Food Services"),
    ("MSW",     "JAB_MSW",                    "high",   "Medical Social Work"),
    ("INFRA",   "JAB_INFRA",                  "high",   "Infrastructure"),
    ("INFO",    "JAB_INFOSTRUKTUR",           "high",   "Infostructure / IT"),
    ("SUV",     "",                           "none",   "Couldn't match. Surveillance? Or AMO Supervision (UNIT_AMO_SUP)?"),
    ("CMX",     "JAB_KOMUNITI_AMBULATORI",    "low",    "Best guess: Community & Ambulatory Services. Confirm."),
    ("PTAR",    "PERP_TAR",                   "high",   "Tun Abdul Razak Medical Library"),
    ("CSSU",    "NUR_CSSU",                   "high",   "Central Sterile Services Unit (subunit of Nursing)"),
    ("CMU",     "UNIT_CASEMIX",               "medium", "Best guess: Casemix Unit. Confirm."),
    # HR (JPSM family)
    ("JPSMADM", "",                           "none",   "HR Admin office — JAB_HRM subunit? Not in pscs_departments yet"),
    ("JPSMGOV", "BAH_GOVERNAN",               "low",    "Possibly Governance & Integrity Division"),
    ("JPSMHRM", "JAB_HRM",                    "high",   "Human Resource Management"),
    ("JPSMHRD", "",                           "none",   "HR Development — not in pscs_departments yet"),
    ("JPSMBDU", "BAH_BIZ_DEV",                "medium", "Business Development Unit"),
    ("JPSMSTL", "UNIT_SPIRITUALITI",          "medium", "Spirituality Unit?"),
    ("JPSMCOUN","UNIT_KAUNSELING",            "high",   "Counselling Unit"),
    ("JPSMUSK", "",                           "none",   "JPSM USK = ? Couldn't decode"),
    # Finance
    ("FIN",     "JAB_KEWANGAN",               "high",   "Department of Finance"),
    ("KW",      "",                           "none",   "Couldn't match. Another finance unit?"),
    # Cross-cutting / org
    ("AMS",     "",                           "none",   "Antimicrobial Stewardship — not in pscs_departments yet"),
    ("RMCQ",    "JAB_QPS",                    "medium", "RMCQ Unit — same as QIPS (JAB_QPS)? Or separate?"),
]

# ----- All 104 pscs_departments rows for the reference sheet -----
PSCS_DEPTS = [
    ("DIR_HOSP",                   "Hospital Director",                                           "Pengarah Hospital",                                                  "directorate", "",                              1),
    ("DIR_MED",                    "Deputy Director Clinical Services (Medicine)",                "Timbalan Pengarah Klinikal (Perubatan)",                             "directorate", "",                              2),
    ("DIR_SURG",                   "Deputy Director Clinical Services (Surgical)",                "Timbalan Pengarah Klinikal (Pembedahan)",                            "directorate", "",                              3),
    ("DIR_PROF",                   "Deputy Director Professional & Operation",                    "Timbalan Pengarah Profesional & Operasi",                            "directorate", "",                              4),
    ("DIR_MGMT",                   "Deputy Director Management",                                  "Timbalan Pengarah Pengurusan",                                       "directorate", "",                              5),
    ("DIR_FIN",                    "Deputy Director Finance",                                     "Timbalan Pengarah Kewangan",                                         "directorate", "",                              6),
    ("PEJ_PENGARAH_HOSPITAL",      "Hospital Director's Office",                                  "Pejabat Pengarah Hospital",                                          "department",  "DIR_HOSP",                      7),
    ("JAB_KOMUNITI_AMBULATORI",    "Department of Community & Ambulatory Services",               "Jabatan Perkhidmatan Komuniti & Ambulatori",                         "department",  "DIR_HOSP",                      8),
    ("JAB_PENYELIDIKAN",           "Department of Research, Industry Linkages & Innovation",      "Jabatan Penyelidikan, Jaringan Industri & Inovasi",                  "department",  "DIR_HOSP",                      9),
    ("JAB_KEJURURAWATAN",          "Department of Nursing",                                       "Jabatan Kejururawatan",                                              "department",  "DIR_HOSP",                     10),
    ("PEJ_UNDANG_UNDANG",          "Legal Office",                                                "Pejabat Undang-Undang",                                              "department",  "DIR_HOSP",                     11),
    ("BAH_PERKHIDMATAN_EKSEKUTIF", "Division of Executive Services",                              "Bahagian Perkhidmatan Eksekutif",                                    "department",  "DIR_HOSP",                     12),
    ("BAH_PENGURUSAN_KLINIKAL",    "Division of Clinical Management",                             "Bahagian Pengurusan Klinikal",                                       "department",  "DIR_HOSP",                     13),
    ("PEJ_TPK_MED",                "Office of Deputy Director Clinical Services (Medicine)",      "Pejabat Timbalan Pengarah Klinikal (Perubatan)",                     "department",  "DIR_MED",                      14),
    ("JAB_PERUBATAN",              "Department of Medicine",                                      "Jabatan Perubatan",                                                  "department",  "DIR_MED",                      15),
    ("JAB_PERUBATAN_KECEMASAN",    "Department of Emergency Medicine",                            "Jabatan Perubatan Kecemasan",                                        "department",  "DIR_MED",                      16),
    ("JAB_PERUBATAN_PRIMER",       "Department of Primary Care Medicine",                         "Jabatan Perubatan Penjagaan Primer",                                 "department",  "DIR_MED",                      17),
    ("JAB_PERUBATAN_PEMULIHAN",    "Department of Rehabilitation Medicine",                       "Jabatan Perubatan Pemulihan",                                        "department",  "DIR_MED",                      18),
    ("JAB_PEDIATRIK",              "Department of Paediatrics",                                   "Jabatan Pediatrik",                                                  "department",  "DIR_MED",                      19),
    ("JAB_PSIKIATRI",              "Department of Psychiatry",                                    "Jabatan Psikiatri",                                                  "department",  "DIR_MED",                      20),
    ("JAB_KESIHATAN_AWAM",         "Department of Public Health",                                 "Jabatan Perubatan Kesihatan Awam",                                   "department",  "DIR_MED",                      21),
    ("JAB_ETIKA_PERUBATAN",        "Department of Medical Ethics & Medical Laws",                 "Jabatan Etika & Undang-Undang Perubatan",                            "department",  "DIR_MED",                      22),
    ("PUS_NEFROLOGI",              "Nephrology Service Centre",                                   "Pusat Perkhidmatan Nefrologi",                                       "department",  "DIR_MED",                      23),
    ("PUS_KARDIOLOGI",             "Cardiology Service Centre",                                   "Pusat Perkhidmatan Kardiologi",                                      "department",  "DIR_MED",                      24),
    ("PUS_RESPIRATORI",            "Respiratory Service Centre",                                  "Pusat Perkhidmatan Respiratori",                                     "department",  "DIR_MED",                      25),
    ("PUS_RAWATAN_HARIAN",         "Day Care Service Centre",                                     "Pusat Perkhidmatan Rawatan Harian",                                  "department",  "DIR_MED",                      26),
    ("PUS_GASTRO_HEPATOLOGI",      "Gastroenterology & Hepatology Service Centre",                "Pusat Perkhidmatan Gastroenterologi & Hepatologi",                   "department",  "DIR_MED",                      27),
    ("UNIT_INFEKSI",               "Infection Prevention & Control Unit",                         "Unit Pencegahan & Kawalan Infeksi",                                  "department",  "DIR_MED",                      28),
    ("PEJ_TPK_SURG",               "Office of Deputy Director Clinical Services (Surgical)",      "Pejabat Timbalan Pengarah Klinikal (Pembedahan)",                    "department",  "DIR_SURG",                     29),
    ("JAB_PEMBEDAHAN",             "Department of Surgery",                                       "Jabatan Pembedahan",                                                 "department",  "DIR_SURG",                     30),
    ("JAB_CVTS",                   "Department of Cardiovascular & Thoracic Surgery",             "Jabatan Kardiovaskular & Pembedahan Torasik",                        "department",  "DIR_SURG",                     31),
    ("JAB_ORTHO",                  "Department of Orthopaedics & Traumatology",                   "Jabatan Ortopedik & Traumatologi",                                   "department",  "DIR_SURG",                     32),
    ("JAB_OFTALMOLOGI",            "Department of Ophthalmology",                                 "Jabatan Oftalmologi",                                                "department",  "DIR_SURG",                     33),
    ("JAB_ORL",                    "Department of Otorhinolaryngology – Head & Neck Surgery",     "Jabatan Otorinolaringologi – Pembedahan Kepala & Leher",             "department",  "DIR_SURG",                     34),
    ("JAB_ANESTESIOLOGI",          "Department of Anaesthesiology & Intensive Care",              "Jabatan Anestesiologi & Rawatan Intensif",                           "department",  "DIR_SURG",                     35),
    ("JAB_OBG",                    "Department of Obstetrics & Gynaecology",                      "Jabatan Obstetrik & Ginekologi",                                     "department",  "DIR_SURG",                     36),
    ("PUS_PLASTIK",                "Plastic Surgery Service Centre",                              "Pusat Perkhidmatan Pembedahan Plastik",                              "department",  "DIR_SURG",                     37),
    ("PUS_OT",                     "Operation Theatre Service Centre",                            "Pusat Perkhidmatan Dewan Bedah",                                     "department",  "DIR_SURG",                     38),
    ("PEJ_TPK_PROF",               "Office of Deputy Director Professional & Operation",          "Pejabat Timbalan Pengarah Profesional & Operasi",                    "department",  "DIR_PROF",                     39),
    ("JAB_RADIOLOGI",              "Department of Radiology",                                     "Jabatan Radiologi",                                                  "department",  "DIR_PROF",                     40),
    ("JAB_MAKMAL",                 "Department of Clinical Diagnostic Laboratory",                "Jabatan Makmal Diagnostik Klinikal",                                 "department",  "DIR_PROF",                     41),
    ("JAB_QPS",                    "Department of Risk Management, Compliance & Quality",         "Jabatan Pengurusan Risiko, Pematuhan & Kualiti",                     "department",  "DIR_PROF",                     42),
    ("JAB_FARMASI",                "Department of Pharmacy",                                      "Jabatan Farmasi",                                                    "department",  "DIR_PROF",                     43),
    ("JAB_MAKLUMAT_PESAKIT",       "Department of Patient Information",                           "Jabatan Maklumat Pesakit",                                           "department",  "DIR_PROF",                     44),
    ("JAB_DIETETIK",               "Department of Dietetics & Food Services",                     "Jabatan Dietetik & Sajian",                                          "department",  "DIR_PROF",                     45),
    ("JAB_MSW",                    "Department of Medical Social Work",                           "Jabatan Kerja Sosial Perubatan",                                     "department",  "DIR_PROF",                     46),
    ("JAB_INFRA",                  "Department of Infrastructure",                                "Jabatan Infrastruktur",                                              "department",  "DIR_PROF",                     47),
    ("JAB_INFOSTRUKTUR",           "Department of Infostructure",                                 "Jabatan Infostruktur",                                               "department",  "DIR_PROF",                     48),
    ("UNIT_AMO_SUP",               "Assistant Medical Officer Supervisory Unit",                  "Unit Penyeliaan Penolong Pegawai Perubatan",                         "department",  "DIR_PROF",                     49),
    ("UNIT_CASEMIX",               "Casemix Unit",                                                "Unit Casemix",                                                       "department",  "DIR_PROF",                     50),
    ("PERP_TAR",                   "Tun Abdul Razak Medical Library",                             "Perpustakaan Perubatan Tun Abdul Razak",                             "department",  "DIR_PROF",                     51),
    ("PEJ_TPK_MGMT",               "Office of Deputy Director Management",                       "Pejabat Timbalan Pengarah Pengurusan",                               "department",  "DIR_MGMT",                     52),
    ("JAB_HRM",                    "Department of Human Resource Management",                    "Jabatan Pengurusan Sumber Manusia",                                  "department",  "DIR_MGMT",                     53),
    ("BAH_GOVERNAN",               "Division of Governance & Integrity",                          "Bahagian Governan & Integriti",                                      "department",  "DIR_MGMT",                     54),
    ("BAH_BIZ_DEV",                "Division of Business Development",                            "Bahagian Pembangunan Perniagaan",                                    "department",  "DIR_MGMT",                     55),
    ("PEJ_POLIS_BANTUAN",          "Auxiliary Police Office",                                     "Pejabat Polis Bantuan",                                              "department",  "DIR_MGMT",                     56),
    ("UNIT_SPIRITUALITI",          "Spirituality Unit",                                           "Unit Spiritualiti",                                                  "department",  "DIR_MGMT",                     57),
    ("UNIT_KAUNSELING",            "Counselling Unit",                                            "Unit Kaunseling",                                                    "department",  "DIR_MGMT",                     58),
    ("JAB_KORPORAT",               "Department of Corporate Communication",                       "Jabatan Komunikasi Korporat",                                        "department",  "DIR_MGMT",                     59),
    ("PEJ_TPK_FIN",                "Office of Deputy Director Finance",                           "Pejabat Timbalan Pengarah Kewangan",                                 "department",  "DIR_FIN",                      60),
    ("JAB_KEWANGAN",               "Department of Finance",                                       "Jabatan Kewangan",                                                   "department",  "DIR_FIN",                      61),
    ("NUR_ADMIN",                  "Nursing Administration Office",                              "Pejabat Pentadbiran Kejururawatan",                                  "subunit",     "JAB_KEJURURAWATAN",             62),
    ("NUR_BED_MGMT",               "Bed Management Unit",                                         "Unit Pengurusan Katil",                                              "subunit",     "JAB_KEJURURAWATAN",             63),
    ("NUR_CSSU",                   "Central Sterile Services Unit",                              "Unit Bekalan Steril",                                                "subunit",     "JAB_KEJURURAWATAN",             64),
    ("EXEC_ADMIN",                 "Executive Services Administration Office",                   "Pejabat Pentadbiran Perkhidmatan Eksekutif",                         "subunit",     "BAH_PERKHIDMATAN_EKSEKUTIF",   65),
    ("EXEC_CLINIC",                "Executive Clinic",                                            "Klinik Eksekutif",                                                   "subunit",     "BAH_PERKHIDMATAN_EKSEKUTIF",   66),
    ("EXEC_WARD",                  "Executive Ward",                                              "Wad Eksekutif",                                                      "subunit",     "BAH_PERKHIDMATAN_EKSEKUTIF",   67),
    ("PERUB_KP",                   "Specialist Clinic (Medicine)",                               "Klinik Pakar Perubatan",                                             "subunit",     "JAB_PERUBATAN",                68),
    ("PERUB_WARD",                 "Ward (Medicine)",                                             "Wad Perubatan",                                                      "subunit",     "JAB_PERUBATAN",                69),
    ("REHAB_KP",                   "Specialist Clinic (Rehabilitation Medicine)",                "Klinik Pakar Perubatan Pemulihan",                                   "subunit",     "JAB_PERUBATAN_PEMULIHAN",      70),
    ("REHAB_WARD",                 "Ward (Rehabilitation Medicine)",                              "Wad Perubatan Pemulihan",                                            "subunit",     "JAB_PERUBATAN_PEMULIHAN",      71),
    ("PED_KP",                     "Specialist Clinic (Paediatrics)",                            "Klinik Pakar Pediatrik",                                             "subunit",     "JAB_PEDIATRIK",                72),
    ("PED_WARD",                   "Ward (Paediatrics)",                                          "Wad Pediatrik",                                                      "subunit",     "JAB_PEDIATRIK",                73),
    ("PED_NICU",                   "Neonatal Intensive Care Unit (NICU)",                        "Unit Rawatan Rapi Neonatal (NICU)",                                  "subunit",     "JAB_PEDIATRIK",                74),
    ("PED_PICU",                   "Paediatric Intensive Care Unit (PICU)",                      "Unit Rawatan Rapi Pediatrik (PICU)",                                 "subunit",     "JAB_PEDIATRIK",                75),
    ("PSY_KP",                     "Specialist Clinic (Psychiatry)",                              "Klinik Pakar Psikiatri",                                             "subunit",     "JAB_PSIKIATRI",                76),
    ("PSY_WARD",                   "Ward (Psychiatry)",                                           "Wad Psikiatri",                                                      "subunit",     "JAB_PSIKIATRI",                77),
    ("CARD_KP",                    "Specialist Clinic (Cardiology)",                              "Klinik Pakar Kardiologi",                                            "subunit",     "PUS_KARDIOLOGI",               78),
    ("CARD_WARD",                  "Ward (Cardiology)",                                           "Wad Kardiologi",                                                     "subunit",     "PUS_KARDIOLOGI",               79),
    ("CARD_CCU",                   "Coronary Care Unit (CCU)",                                    "Unit Penjagaan Koronari (CCU)",                                      "subunit",     "PUS_KARDIOLOGI",               80),
    ("CARD_CRW",                   "Cardiac Rehabilitation Ward (CRW)",                          "Wad Pemulihan Kardio (CRW)",                                         "subunit",     "PUS_KARDIOLOGI",               81),
    ("DAY_HARIAN",                 "Daycare Unit",                                                "Unit Rawatan Harian",                                                "subunit",     "PUS_RAWATAN_HARIAN",           82),
    ("DAY_OT",                     "Operation Theatre Unit (Day Care)",                          "Unit Dewan Bedah (Rawatan Harian)",                                  "subunit",     "PUS_RAWATAN_HARIAN",           83),
    ("DAY_ENDO",                   "Endoscopy Unit",                                              "Unit Endoskopi",                                                     "subunit",     "PUS_RAWATAN_HARIAN",           84),
    ("SURG_KP",                    "Specialist Clinic (Surgery)",                                 "Klinik Pakar Pembedahan",                                            "subunit",     "JAB_PEMBEDAHAN",               85),
    ("SURG_WARD",                  "Ward (Surgery)",                                              "Wad Pembedahan",                                                     "subunit",     "JAB_PEMBEDAHAN",               86),
    ("CVTS_KP",                    "Specialist Clinic (Cardiovascular & Thoracic Surgery)",      "Klinik Pakar Kardiovaskular & Pembedahan Torasik",                   "subunit",     "JAB_CVTS",                     87),
    ("CVTS_WARD",                  "Ward (Cardiovascular & Thoracic Surgery)",                   "Wad Kardiovaskular & Pembedahan Torasik",                            "subunit",     "JAB_CVTS",                     88),
    ("ORTHO_KP",                   "Specialist Clinic (Orthopaedics & Traumatology)",            "Klinik Pakar Ortopedik & Traumatologi",                              "subunit",     "JAB_ORTHO",                    89),
    ("ORTHO_WARD",                 "Ward (Orthopaedics & Traumatology)",                         "Wad Ortopedik & Traumatologi",                                       "subunit",     "JAB_ORTHO",                    90),
    ("OFTAL_KP",                   "Specialist Clinic (Ophthalmology)",                          "Klinik Pakar Oftalmologi",                                           "subunit",     "JAB_OFTALMOLOGI",              91),
    ("OFTAL_WARD",                 "Ward (Ophthalmology)",                                        "Wad Oftalmologi",                                                    "subunit",     "JAB_OFTALMOLOGI",              92),
    ("ORL_KP",                     "Specialist Clinic (ORL – Head & Neck)",                       "Klinik Pakar Otorinolaringologi – Pembedahan Kepala & Leher",        "subunit",     "JAB_ORL",                      93),
    ("ORL_WARD",                   "Ward (ORL – Head & Neck)",                                    "Wad Otorinolaringologi – Pembedahan Kepala & Leher",                 "subunit",     "JAB_ORL",                      94),
    ("ANES_KP",                    "Specialist Clinic (Anaesthesiology & Intensive Care)",       "Klinik Pakar Anestesiologi & Rawatan Intensif",                      "subunit",     "JAB_ANESTESIOLOGI",            95),
    ("ANES_ICU_HASA",              "Intensive Care Unit (ICU – HASA)",                            "Unit Rawatan Rapi (ICU – HASA)",                                     "subunit",     "JAB_ANESTESIOLOGI",            96),
    ("ANES_ICU_PPUITM",            "Intensive Care Unit (ICU – PPUiTM)",                         "Unit Rawatan Rapi (ICU – PPUiTM)",                                   "subunit",     "JAB_ANESTESIOLOGI",            97),
    ("ANES_CICU_PPUITM",           "Cardiothoracic Intensive Care Unit (CICU – PPUiTM)",         "Unit Rawatan Rapi Kardiotorasik (CICU – PPUiTM)",                    "subunit",     "JAB_ANESTESIOLOGI",            98),
    ("OBG_KP",                     "Specialist Clinic (Obstetrics & Gynaecology)",                "Klinik Pakar Obstetrik & Ginekologi",                                "subunit",     "JAB_OBG",                      99),
    ("OBG_WARD",                   "Ward (Obstetrics & Gynaecology)",                             "Wad Obstetrik & Ginekologi",                                         "subunit",     "JAB_OBG",                     100),
    ("OBG_DELIVERY",               "Labour Room (Dewan Bersalin)",                                "Dewan Bersalin",                                                     "subunit",     "JAB_OBG",                     101),
    ("PLAS_KP",                    "Specialist Clinic (Plastic Surgery)",                         "Klinik Pakar Pembedahan Plastik",                                    "subunit",     "PUS_PLASTIK",                 102),
    ("PLAS_WARD",                  "Ward (Plastic Surgery)",                                      "Wad Pembedahan Plastik",                                             "subunit",     "PUS_PLASTIK",                 103),
    ("AMO_SUP_ADMIN",              "AMO Supervisory Administration Office",                       "Pejabat Pentadbiran Unit Penyeliaan Penolong Pegawai Perubatan",     "subunit",     "UNIT_AMO_SUP",                104),
]

wb = Workbook()

# ===== Sheet 1: Mapping =====
ws = wb.active
ws.title = "Mapping"
ws.append(["risk_code", "your_pscs_code (FILL IN)", "my_suggestion", "confidence", "notes / what it likely is"])

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", start_color="185FA5")
input_fill  = PatternFill("solid", start_color="FFF7CC")
none_fill   = PatternFill("solid", start_color="FBEAEA")
high_fill   = PatternFill("solid", start_color="E5F3DC")
medium_fill = PatternFill("solid", start_color="FAEEDA")

# Headers
for col_idx, _ in enumerate(["risk_code", "your_pscs_code (FILL IN)", "my_suggestion", "confidence", "notes / what it likely is"], start=1):
    c = ws.cell(row=1, column=col_idx)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="left", vertical="center")

# Data
for i, (risk_code, suggested, confidence, notes) in enumerate(MAPPING, start=2):
    ws.cell(row=i, column=1, value=risk_code)
    your_cell = ws.cell(row=i, column=2, value=suggested if confidence == "high" else "")
    your_cell.fill = input_fill
    ws.cell(row=i, column=3, value=suggested)
    conf_cell = ws.cell(row=i, column=4, value=confidence)
    if confidence == "high":
        conf_cell.fill = high_fill
    elif confidence == "medium":
        conf_cell.fill = medium_fill
    else:
        conf_cell.fill = none_fill
    ws.cell(row=i, column=5, value=notes)

# Column widths
for col, width in [("A", 14), ("B", 32), ("C", 32), ("D", 12), ("E", 70)]:
    ws.column_dimensions[col].width = width
ws.freeze_panes = "A2"

# ===== Sheet 2: PSCS_Reference =====
ref = wb.create_sheet("PSCS_Reference")
ref.append(["code", "name_en", "name_ms", "kind", "parent_code", "sort_order"])
for col_idx in range(1, 7):
    c = ref.cell(row=1, column=col_idx)
    c.font = header_font
    c.fill = header_fill
for row in PSCS_DEPTS:
    ref.append(list(row))
for col, width in [("A", 30), ("B", 60), ("C", 60), ("D", 14), ("E", 30), ("F", 10)]:
    ref.column_dimensions[col].width = width
ref.freeze_panes = "A2"

# ===== Sheet 3: README =====
readme = wb.create_sheet("README", 0)
readme.append(["HASA Risk Module — risk_code Mapping Worksheet"])
readme.append([])
readme.append(["How to use:"])
readme.append(["1. Open the Mapping sheet."])
readme.append(["2. For each risk_code, fill in the YOUR_PSCS_CODE column (yellow)."])
readme.append(["3. Use the PSCS_Reference sheet to look up codes by name (Cmd+F)."])
readme.append(["4. If a Risk code doesn't exist in pscs_departments yet, type ADD_NEW in the cell + describe in notes."])
readme.append(["5. If a Risk code should be dropped entirely, type SKIP."])
readme.append(["6. Save the file and hand it back. I'll run the UPDATE statements against pscs_departments.risk_code."])
readme.append([])
readme.append(["Pre-filled highlights:"])
readme.append(["  - High-confidence rows (green) are pre-filled in your column. Review and edit if needed."])
readme.append(["  - Medium-confidence rows (amber) have my best guess in the suggestion column but YOUR cell is blank."])
readme.append(["  - No-match rows (red) need your input. Read the notes for what I think it might be."])
readme.append([])
readme.append(["Counts:"])
high_n   = sum(1 for r in MAPPING if r[2] == "high")
medium_n = sum(1 for r in MAPPING if r[2] == "medium")
none_n   = sum(1 for r in MAPPING if r[2] == "none")
readme.append([f"  {high_n} high-confidence (pre-filled)"])
readme.append([f"  {medium_n} medium-confidence (review my guess)"])
readme.append([f"  {none_n} no match (please specify)"])
readme.append([f"  {len(MAPPING)} total risk codes"])
readme.cell(row=1, column=1).font = Font(bold=True, size=14)
readme.column_dimensions["A"].width = 120

import sys
out = sys.argv[1] if len(sys.argv) > 1 else "Risk_Code_Mapping.xlsx"
wb.save(out)
print(f"Saved to {out}")
